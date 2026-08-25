from openpyxl import load_workbook

from ragkb.models import ParsedDocument, TableData
from ragkb.parsers.base import DocumentParser, cell_to_str


class ExcelParser(DocumentParser):
    """Excel：每个 sheet 独立成表，保留表头与行列结构。

    假设第一个非空行为表头（has_headers=True 时的默认行为）；
    无表头的数据传 has_headers=False，此时所有行都进 rows，
    headers 为空列表。
    """

    def parse(self, path, doc_id, source, department="", version="",
              effective_date="", has_headers=True):
        wb = load_workbook(path, data_only=True)
        tables = []
        for ws in wb.worksheets:
            rows = [[cell_to_str(c.value) for c in row]
                    for row in ws.iter_rows()]
            rows = [r for r in rows if any(v != "N/A" for v in r)]
            if not rows:
                continue
            headers = rows[0] if has_headers else []
            data_rows = rows[1:] if has_headers else rows
            tables.append(TableData(
                table_id=f"{source}-{ws.title}",
                name=ws.title,
                headers=headers,
                rows=data_rows,
                source=f"{source}:{ws.title}",
                doc_id=doc_id,
            ))
        return ParsedDocument(
            doc_id=doc_id, doc_type="excel", source=source, text="",
            tables=tables, images=[],
            department=department, version=version, effective_date=effective_date,
        )
